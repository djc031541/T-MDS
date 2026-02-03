from __future__ import annotations

import io
import os
import re
from dataclasses import dataclass
from datetime import datetime, timedelta
import hashlib
import html
from pathlib import Path
import hmac
from typing import Any, Iterable

import pandas as pd
import streamlit as st

try:
    import matplotlib.pyplot as plt
    import matplotlib.font_manager as fm

    MATPLOTLIB_AVAILABLE = True
except Exception:
    MATPLOTLIB_AVAILABLE = False

try:
    import google.generativeai as genai

    GEMINI_AVAILABLE = True
except Exception:
    GEMINI_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False


APP_TITLE = "T'way MDS - Maintenance Document System"
DEFAULT_APP_PASSWORD = "twatmcc2026"


FLEET_MAP: dict[str, list[str]] = {
    "B737-800": [
        "HL8000",
        "HL8098",
        "HL8047",
        "HL8306",
        "HL8056",
        "HL8565",
        "HL8067",
        "HL8030",
        "HL8069",
        "HL8323",
        "HL8070",
        "HL8235",
        "HL8086",
        "HL8737",
        "HL8095",
        "HL8220",
        "HL8233",
        "HL8300",
        "HL8324",
        "HL8326",
        "HL8327",
        "HL8329",
        "HL8354",
        "HL8363",
        "HL8373",
        "HL8379",
        "HL8547",
        "HL8564",
        "HL8378",
    ],
    "B737-8": ["HL8513", "HL8514", "HL8580", "HL8581"],
    "B777-300ER": ["HL8706", "HL8707"],
    "A330-200": ["HL8211", "HL8212", "HL8227", "HL8228", "HL8276", "HL8708"],
    "A330-300": ["HL8500", "HL8501", "HL8502", "HL8560", "HL8561"],
}

EXCLUDED_ATA = {"32", "33", "34"}  # TIRE/BRAKE, LIGHT, GPS


@dataclass(frozen=True)
class SearchFilters:
    actype: str
    reg: str
    wo: str
    desc: str
    action: str
    col_n: str
    ata: str
    date_from: str  # YYYY-MM-DD or empty
    date_to: str  # YYYY-MM-DD or empty


def _clean_col_name(col: object) -> str:
    s = "" if col is None else str(col)
    s = s.replace("\ufeff", "").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [_clean_col_name(c) for c in df.columns]
    return df


def _try_read_csv_from_bytes(data: bytes) -> pd.DataFrame:
    last_err: Exception | None = None
    for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            # Keep everything as strings (the original Tkinter app uses dtype=str).
            try:
                return pd.read_csv(io.BytesIO(data), encoding=enc, dtype=str, engine="pyarrow", low_memory=False)
            except Exception:
                return pd.read_csv(io.BytesIO(data), encoding=enc, dtype=str, low_memory=False)
        except Exception as e:  # noqa: BLE001
            last_err = e
    raise RuntimeError(f"CSV 읽기 실패: {last_err}") from last_err


def _read_dataframe(file_name: str, file_bytes: bytes) -> pd.DataFrame:
    name = (file_name or "").lower()
    if name.endswith(".csv"):
        return _try_read_csv_from_bytes(file_bytes)
    if name.endswith(".xlsx") or name.endswith(".xls"):
        # Use openpyxl. Convert to string-ish like the original.
        df = pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl")
        df = df.fillna("").astype(str)
        return df
    raise ValueError("지원하지 않는 파일 형식입니다. CSV 또는 XLSX를 업로드해 주세요.")


@st.cache_data(show_spinner=False)
def load_data_from_upload(file_name: str, file_bytes: bytes) -> pd.DataFrame:
    df = _read_dataframe(file_name, file_bytes)
    df = df.fillna("")
    return _clean_columns(df)


@st.cache_data(show_spinner=False)
def load_data_from_path(path: str) -> pd.DataFrame:
    if not os.path.exists(path):
        raise FileNotFoundError(path)

    lower = path.lower()
    if lower.endswith(".csv"):
        # Preserve original encoding fallback behavior.
        for enc in ("cp949", "utf-8-sig", "utf-8"):
            try:
                df = pd.read_csv(path, encoding=enc, dtype=str, on_bad_lines="skip", low_memory=False)
                return _clean_columns(df.fillna(""))
            except Exception:  # noqa: BLE001
                continue
        df = pd.read_csv(path, dtype=str, on_bad_lines="skip", low_memory=False)
        return _clean_columns(df.fillna(""))

    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        df = pd.read_excel(path, engine="openpyxl")
        df = df.fillna("").astype(str)
        return _clean_columns(df)

    raise ValueError("지원하지 않는 파일 형식입니다. CSV 또는 XLSX 경로를 입력해 주세요.")


def determine_ac_type(reg: str) -> str:
    if not reg:
        return ""
    for ac, regs in FLEET_MAP.items():
        if any(r in str(reg) for r in regs):
            return ac
    return ""


def parse_date(s: Any) -> datetime | None:
    if s is None or (isinstance(s, float) and pd.isna(s)) or pd.isna(s):
        return None
    s = str(s).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:10], fmt)
        except Exception:  # noqa: BLE001
            pass
    return None


def format_date(val: Any) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)) or pd.isna(val):
        return ""
    s = str(val).strip()
    if "@" in s or len(s) < 6:
        return ""
    parsed = parse_date(s)
    if parsed:
        return parsed.strftime("%Y-%m-%d")
    return s[:10] if len(s) >= 10 else s


def extract_symptom_key(desc: Any, ata: Any) -> str:
    d = str(desc).upper().strip()
    ata_str = str(ata).strip()

    # ATA 25: seat-position-like defects should not be collapsed by position tokens.
    if ata_str == "25":
        d = re.sub(r"\b(LH|RH|LEFT|RIGHT)\b", "", d)
        d = re.sub(r"\s+", " ", d).strip()
        return d

    d = re.sub(r"\b\d{1,2}[A-KL-Z]\b", "", d)
    d = re.sub(
        r"\b(LH|RH|LEFT|RIGHT|FWD|AFT|FORWARD|REAR|UPPER|LOWER|INBD|OUTBD|INBOARD|OUTBOARD)\b",
        "",
        d,
    )
    d = re.sub(r"\b(NO\.?\s*\d+|#\d+|POS\.?\s*\d+|ZONE\.?\s*\d+|DOOR\.?\s*\d+|LAV\.?\s*\d+|PSU\.?\s*\d+)\b", "", d)
    d = re.sub(r"\s+", " ", d).strip()
    return d


def is_mel_cdl_repetitive(desc: Any, action: Any) -> bool:
    combined = (str(desc) + " " + str(action)).upper()
    patterns = [
        r"\bMEL\b",
        r"\bCDL\b",
        r"MEL\s*ITEM",
        r"CDL\s*ITEM",
        r"MEL\s*#",
        r"CDL\s*#",
        r"REPETITIVE\s*INSP",
        r"REPETITIVE\s*CHECK",
        r"REP\s*INSP",
        r"R/I\s*DUE",
        r"REPEAT\s*INSPECTION",
        r"RECURRING\s*INSP",
    ]
    return any(re.search(p, combined) for p in patterns)


def is_cosmetic_defect(desc: Any) -> bool:
    keywords = [
        "paint",
        "scratch",
        "dent",
        "nick",
        "scuff",
        "peel",
        "peeling",
        "chip",
        "chipped",
        "worn",
        "dirty",
        "clean",
        "stain",
        "discolor",
        "corrosion",
        "erosion",
        "gouge",
    ]
    d = str(desc).lower()
    return any(kw in d for kw in keywords)


def build_mds_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Reproduce the original Tkinter app's column mapping behavior.
    Produces:
    - My_WO, My_Order_No, My_Desc, My_Col_N, My_Ref_Doc, My_ATA
    - My_Date, My_Action, My_Reg, My_AC_Str, Parsed_Date
    """
    df = df.copy()
    df = df.fillna("")
    n = df.shape[1]

    df["My_WO"] = df.iloc[:, 2] if n > 2 else ""
    df["My_Order_No"] = df.iloc[:, 3] if n > 3 else ""
    df["My_Desc"] = df.iloc[:, 5] if n > 5 else ""
    df["My_Col_N"] = df.iloc[:, 13] if n > 13 else ""
    df["My_Ref_Doc"] = df.iloc[:, 73] if n > 73 else ""
    df["My_ATA"] = df["My_Ref_Doc"].astype(str).str.slice(0, 2)

    date_col_idx: int | None = None
    for i, col in enumerate(df.columns):
        c = str(col)
        if "date" in c.lower() or "날짜" in c:
            date_col_idx = i
            break
    if date_col_idx is None:
        date_col_idx = 1 if n > 1 else 0
    df["My_Date"] = df.iloc[:, date_col_idx] if n > date_col_idx else ""

    action_col = next((c for c in df.columns if "Action" in str(c)), None)
    df["My_Action"] = df[action_col] if action_col in df.columns else ""

    reg_col = next((c for c in df.columns if ("Reg" in str(c)) or ("Loc" in str(c))), None)
    df["My_Reg"] = df[reg_col].astype(str).str.strip() if reg_col in df.columns else ""

    df["My_AC_Str"] = df["My_Reg"].apply(determine_ac_type)

    # Fast path: vectorized datetime parsing; fallback to the legacy parser for leftovers.
    s_date = df["My_Date"].astype(str).str.strip()
    parsed = pd.to_datetime(s_date, errors="coerce")
    missing = parsed.isna() & s_date.ne("")
    if bool(missing.any()):
        parsed2 = s_date[missing].apply(parse_date)
        parsed.loc[missing] = pd.to_datetime(parsed2, errors="coerce")
    df["Parsed_Date"] = parsed

    # Precompute a display-ready date string so we don't reformat on every rerun.
    df["Display_Date"] = df["Parsed_Date"].dt.strftime("%Y-%m-%d")
    # For unparsable dates, keep a compact best-effort preview.
    df.loc[df["Display_Date"].isna(), "Display_Date"] = s_date.loc[df["Display_Date"].isna()].str.slice(0, 10)

    return df


def analyze_repetitive_defects(df_with_mds_cols: pd.DataFrame) -> list[dict[str, Any]]:
    if df_with_mds_cols.empty:
        return []

    today = datetime.now()
    four_weeks_ago = today - timedelta(days=28)

    df = df_with_mds_cols[df_with_mds_cols["Parsed_Date"].notna()].copy()
    df = df[df["Parsed_Date"] >= four_weeks_ago]
    if df.empty:
        return []

    df = df[~df["My_ATA"].isin(EXCLUDED_ATA)]
    df = df[~df["My_Desc"].apply(is_cosmetic_defect)]
    df = df[~df.apply(lambda r: is_mel_cdl_repetitive(r["My_Desc"], r["My_Action"]), axis=1)]
    if df.empty:
        return []

    df["Symptom_Key"] = df.apply(lambda r: extract_symptom_key(r["My_Desc"], r["My_ATA"]), axis=1)

    result: list[dict[str, Any]] = []
    for (reg, ata, symptom), group in df.groupby(["My_Reg", "My_ATA", "Symptom_Key"]):
        if not reg or not ata or not symptom:
            continue
        if len(str(symptom)) < 3:
            continue

        c_1w = len(group[group["Parsed_Date"] >= (today - timedelta(days=7))])
        c_2w = len(group[group["Parsed_Date"] >= (today - timedelta(days=14))])
        c_4w = len(group[group["Parsed_Date"] >= four_weeks_ago])

        is_repetitive = (c_1w >= 3) or (c_2w >= 5) or (c_4w >= 8)
        if not is_repetitive:
            continue

        if c_1w >= 3:
            criteria = f"1주 {c_1w}회"
            severity = 3
        elif c_2w >= 5:
            criteria = f"2주 {c_2w}회"
            severity = 2
        else:
            criteria = f"4주 {c_4w}회"
            severity = 1

        sample_desc = str(group.iloc[0]["My_Desc"])[:100]
        result.append(
            {
                "status": "🔴 반복결함",
                "reg": reg,
                "ac_type": determine_ac_type(reg),
                "ata": ata,
                "symptom": str(symptom)[:50],
                "count_1w": c_1w,
                "count_2w": c_2w,
                "count_4w": c_4w,
                "criteria": criteria,
                "severity": severity,
                "description": sample_desc,
                "dates": group["Parsed_Date"].dt.strftime("%Y-%m-%d").tolist(),
            }
        )

    result.sort(key=lambda x: (x["severity"], x["count_1w"], x["count_2w"], x["count_4w"]), reverse=True)
    return result


def apply_filters(df: pd.DataFrame, f: SearchFilters) -> pd.DataFrame:
    res = df

    if f.actype:
        if f.actype == "Boeing 공용":
            res = res[res["My_AC_Str"].astype(str).str.startswith("B", na=False)]
        elif f.actype == "Airbus 공용":
            res = res[res["My_AC_Str"].astype(str).str.startswith("A", na=False)]
        else:
            res = res[res["My_AC_Str"] == f.actype]

    if f.reg:
        res = res[res["My_Reg"].astype(str).str.contains(f.reg, case=False, na=False)]
    if f.wo:
        res = res[res["My_WO"].astype(str).str.contains(f.wo, case=False, na=False)]
    if f.desc:
        res = res[res["My_Desc"].astype(str).str.contains(f.desc, case=False, na=False)]
    if f.action:
        res = res[res["My_Action"].astype(str).str.contains(f.action, case=False, na=False)]
    if f.col_n:
        res = res[res["My_Col_N"].astype(str).str.contains(f.col_n, case=False, na=False)]
    if f.ata:
        res = res[res["My_ATA"].astype(str) == f.ata]

    if f.date_from:
        try:
            from_dt = datetime.strptime(f.date_from, "%Y-%m-%d")
            res = res[(res["Parsed_Date"].isna()) | (res["Parsed_Date"] >= from_dt)]
        except Exception:  # noqa: BLE001
            pass
    if f.date_to:
        try:
            to_dt = datetime.strptime(f.date_to, "%Y-%m-%d") + timedelta(days=1)
            res = res[(res["Parsed_Date"].isna()) | (res["Parsed_Date"] < to_dt)]
        except Exception:  # noqa: BLE001
            pass

    return res


STOPWORDS = {
    # English common
    "the",
    "and",
    "for",
    "with",
    "from",
    "this",
    "that",
    "are",
    "was",
    "were",
    "due",
    "done",
    "into",
    "upon",
    "area",
    "check",
    "inspection",
    "inspected",
    "insp",
    "replaced",
    "replace",
    "replace",
    "remove",
    "removed",
    "install",
    "installed",
    "change",
    "changed",
    "corrected",
    "correction",
    "serviceable",
    "per",
    "performed",
    "repair",
    "repaired",
    "test",
    "tested",
    "adjust",
    "adjusted",
    "verified",
    "verify",
    "found",
    "found",
    "by",
    "to",
    "in",
    "out",
    "no",
    "not",
    "ok",
    "nml",
    "iaw",
    "amm",
    "ref",
    "log",
    "unit",
    "assy",
    "s/n",
    "pn",
    "p/n",
    # Korean common
    "조치",
    "조치내용",
    "정상",
    "확인",
    "점검",
    "교체",
    "교환",
    "수리",
    "작업",
    "실시",
    "완료",
    "대체",
    "수행",
    "이상",
    "현상",
    "발생",
    "불량",
    "누유",
    "누설",
    "결함",
}


def extract_keywords(series: pd.Series) -> pd.Series:
    """
    Extract keyword tokens from a text series.
    - Korean/English/Numbers tokens
    - length >= 3
    - drop stopwords and pure digits
    Returns a Series of tokens (lowercased).
    """
    s = series.fillna("").astype(str).str.lower()
    tokens = s.str.findall(r"[0-9A-Za-z가-힣]+").explode()
    if tokens.empty:
        return tokens
    tokens = tokens[tokens.str.len() >= 3]
    tokens = tokens[~tokens.str.fullmatch(r"\d+")]
    tokens = tokens[~tokens.isin(STOPWORDS)]
    return tokens


def pareto_df(counts: pd.Series, label: str = "항목") -> pd.DataFrame:
    if counts is None or counts.empty:
        return pd.DataFrame(columns=[label, "count", "pct", "cum_pct"])
    df = counts.reset_index()
    df.columns = [label, "count"]
    total = df["count"].sum()
    df["pct"] = (df["count"] / total) * 100.0
    df["cum_pct"] = df["pct"].cumsum()
    return df


def _plot_pareto(counts: pd.Series, title: str, xlabel: str) -> "plt.Figure":
    # Build a Pareto chart: bars + cumulative line
    df = pareto_df(counts, label=xlabel)
    fig = plt.figure(figsize=(12, 5))
    ax = fig.add_subplot(1, 1, 1)
    ax.bar(df[xlabel].astype(str), df["count"], color="#e94560")
    ax.set_title(title)
    ax.set_xlabel(xlabel)
    ax.set_ylabel("건수")
    ax.tick_params(axis="x", rotation=45, labelsize=9)

    ax2 = ax.twinx()
    ax2.plot(df[xlabel].astype(str), df["cum_pct"], color="#ffd166", marker="o", linewidth=2)
    ax2.set_ylabel("누적비율(%)")
    ax2.set_ylim(0, 110)
    fig.tight_layout()
    return fig


def setup_korean_font() -> None:
    if not MATPLOTLIB_AVAILABLE:
        return
    # Prefer bundled font for consistent Korean rendering on Streamlit Cloud/Linux.
    font_path = Path(__file__).parent / "assets" / "fonts" / "NanumGothic-Regular.ttf"
    if font_path.exists() and font_path.stat().st_size > 500_000:
        try:
            fm.fontManager.addfont(str(font_path))
            font_prop = fm.FontProperties(fname=str(font_path))
            plt.rcParams["font.family"] = font_prop.get_name()
        except Exception:
            # If font load fails, fall back to system fonts without crashing the app.
            plt.rcParams["font.family"] = ["AppleGothic", "Malgun Gothic", "NanumGothic", "DejaVu Sans"]
    else:
        # Fallback to common system fonts.
        plt.rcParams["font.family"] = ["AppleGothic", "Malgun Gothic", "NanumGothic", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False


def require_login() -> None:
    """
    Simple password gate for Streamlit Cloud.
    Uses st.secrets["APP_PASSWORD"] if set; otherwise falls back to DEFAULT_APP_PASSWORD.
    """
    if st.session_state.get("authenticated"):
        return

    try:
        secret_pw = st.secrets.get("APP_PASSWORD", "")
    except Exception:  # noqa: BLE001
        secret_pw = ""
    if not secret_pw:
        secret_pw = DEFAULT_APP_PASSWORD

    st.title(APP_TITLE)
    st.markdown("접속을 위해 비밀번호가 필요합니다.")

    def _check_password() -> None:
        if hmac.compare_digest(st.session_state.get("password_input", ""), secret_pw):
            st.session_state["authenticated"] = True
            st.session_state.pop("password_input", None)
        else:
            st.session_state["authenticated"] = False

    st.text_input("비밀번호", type="password", key="password_input", on_change=_check_password)
    if st.session_state.get("authenticated") is False:
        st.error("비밀번호가 틀렸습니다.")

    st.stop()

def df_to_xlsx_bytes(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
    """
    Return an XLSX file as bytes. Safe for Streamlit download_button.
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return buf.getvalue()


def _download_excel_bytes(df: pd.DataFrame, filters: SearchFilters) -> bytes:
    """
    Streamlit-friendly equivalent of the Tkinter app's styled Excel export.
    """
    if not OPENPYXL_AVAILABLE:
        return df_to_xlsx_bytes(df, sheet_name="검색결과")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "검색결과"

    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin", color="444444"),
        right=Side(style="thin", color="444444"),
        top=Side(style="thin", color="444444"),
        bottom=Side(style="thin", color="444444"),
    )

    export_cols = ["My_AC_Str", "My_Reg", "My_WO", "My_Order_No", "My_Desc", "My_Action", "My_Col_N", "My_ATA", "My_Ref_Doc", "My_Date"]
    headers = ["AC Type", "Reg No", "Tway W/O", "Order No.", "Description", "Corrective Action", "Col N", "ATA", "Ref Doc", "Date"]

    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(export_cols, 1):
            value = row.get(col_name, "") if col_name in df.columns else ""
            if col_name == "My_Date":
                value = format_date(value)
            cell = ws1.cell(row=row_idx, column=col_idx, value=str(value))
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")

    col_widths = [12, 10, 15, 12, 50, 45, 10, 6, 15, 12]
    for i, width in enumerate(col_widths, 1):
        ws1.column_dimensions[chr(64 + i)].width = width

    # ATA stats sheet (+ chart)
    ws2 = wb.create_sheet("ATA통계")
    ata_stats = df.groupby("My_ATA").size().reset_index(name="건수")
    ata_stats = ata_stats.sort_values("건수", ascending=False)

    ws2.append(["ATA", "건수"])
    for _, r in ata_stats.iterrows():
        ws2.append([str(r["My_ATA"]), int(r["건수"])])

    for c in ws2[1]:
        c.fill = header_fill
        c.font = header_font
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")

    chart = BarChart()
    chart.title = "ATA별 결함 건수"
    chart.y_axis.title = "건수"
    chart.x_axis.title = "ATA"

    data_ref = Reference(ws2, min_col=2, min_row=1, max_row=1 + len(ata_stats))
    cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=1 + len(ata_stats))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 12
    chart.width = 24
    ws2.add_chart(chart, "D2")

    # Summary sheet
    ws3 = wb.create_sheet("요약정보")
    ws3.append(["항목", "값"])
    for c in ws3[1]:
        c.fill = header_fill
        c.font = header_font
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")

    summary = [
        ("AC Type", filters.actype or "전체"),
        ("Reg No", filters.reg or "전체"),
        ("Work Order", filters.wo or "전체"),
        ("Description", filters.desc or "전체"),
        ("Corrective Action", filters.action or "전체"),
        ("Column N", filters.col_n or "전체"),
        ("ATA", filters.ata or "전체"),
        ("기간", f"{filters.date_from or '-'} ~ {filters.date_to or '-'}"),
        ("결과 건수", f"{len(df):,}"),
    ]
    for k, v in summary:
        ws3.append([k, v])
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _init_state() -> None:
    defaults = {
        "filters": SearchFilters(
            actype="",
            reg="",
            wo="",
            desc="",
            action="",
            col_n="",
            ata="",
            date_from="",
            date_to="",
        ),
        "current_result": None,
        "gemini_messages": [],
        "gemini_model": "gemini-2.0-flash",
        "data_key": None,
        "raw_df": None,
        "mds_df": None,
        "repetitive": None,
        "filtered_df": None,
        "filtered_for": None,  # tuple(data_key, filters)
        "selected_row": None,
        "stats_ready": False,
        "upload_name": None,
        "upload_size": None,
        "download_ready_for": None,
        "download_xlsx": None,
        "download_csv": None,
        "stats_key": None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def _set_quick_range(days: int) -> None:
    today = datetime.now().date()
    st.session_state["filters"] = SearchFilters(
        **{**st.session_state["filters"].__dict__, "date_from": (today - timedelta(days=days)).strftime("%Y-%m-%d"), "date_to": today.strftime("%Y-%m-%d")}
    )


def _clear_range() -> None:
    st.session_state["filters"] = SearchFilters(**{**st.session_state["filters"].__dict__, "date_from": "", "date_to": ""})


def _render_sidebar_filters() -> None:
    f: SearchFilters = st.session_state["filters"]

    st.sidebar.header("🔍 검색 조건")

    cols = st.sidebar.columns(5)
    if cols[0].button("오늘"):
        _set_quick_range(0)
    if cols[1].button("1주"):
        _set_quick_range(7)
    if cols[2].button("2주"):
        _set_quick_range(14)
    if cols[3].button("4주"):
        _set_quick_range(28)
    if cols[4].button("전체"):
        _clear_range()

    with st.sidebar.form("filters_form", clear_on_submit=False):
        actype = st.selectbox("AC Type", options=["", "Boeing 공용", "Airbus 공용", *list(FLEET_MAP.keys())], index=0)
        reg = st.text_input("Reg No", value=f.reg)
        wo = st.text_input("Work Order", value=f.wo)
        desc = st.text_input("Description", value=f.desc)
        action = st.text_input("Corrective Action", value=f.action)
        col_n = st.text_input("Column N", value=f.col_n)
        ata = st.text_input("ATA", value=f.ata)

        c1, c2 = st.columns(2)
        date_from = c1.text_input("Date From (YYYY-MM-DD)", value=f.date_from)
        date_to = c2.text_input("Date To (YYYY-MM-DD)", value=f.date_to)

        c3, c4 = st.columns(2)
        do_search = c3.form_submit_button("검색")
        do_reset = c4.form_submit_button("초기화")

    if do_reset:
        st.session_state["filters"] = SearchFilters(
            actype="",
            reg="",
            wo="",
            desc="",
            action="",
            col_n="",
            ata="",
            date_from="",
            date_to="",
        )
        st.session_state["current_result"] = None
        st.session_state["filtered_df"] = None
        st.session_state["filtered_for"] = None
        st.rerun()

    if do_search:
        st.session_state["filters"] = SearchFilters(
            actype=actype,
            reg=reg.strip(),
            wo=wo.strip(),
            desc=desc.strip(),
            action=action.strip(),
            col_n=col_n.strip(),
            ata=ata.strip(),
            date_from=date_from.strip(),
            date_to=date_to.strip(),
        )
        # Invalidate cached filtering for the new criteria.
        st.session_state["filtered_df"] = None
        st.session_state["filtered_for"] = None
        st.rerun()

    st.sidebar.divider()
    st.sidebar.header("📂 데이터")
    st.sidebar.caption("Streamlit Cloud에서는 로컬 경로를 직접 읽을 수 없어서 업로드 방식이 가장 안전합니다.")


def _load_ui() -> pd.DataFrame | None:
    local_path = st.sidebar.text_input("로컬 파일 경로(옵션)", value="", placeholder="/path/to/tway_master_data.csv")
    uploaded = st.sidebar.file_uploader("CSV 또는 XLSX 업로드", type=["csv", "xlsx", "xls"])

    if local_path:
        try:
            with st.sidebar.status("로컬 파일 읽는 중...", expanded=False):
                df = load_data_from_path(local_path)
            st_info = os.stat(local_path)
            key = f"path:{local_path}:{int(st_info.st_mtime)}:{st_info.st_size}"
            st.session_state["data_key"] = key
            st.session_state["raw_df"] = df
            st.session_state["upload_name"] = None
            st.session_state["upload_size"] = None
            st.sidebar.success(f"로컬 파일 로드: {os.path.basename(local_path)}")
            return df
        except Exception as e:  # noqa: BLE001
            st.sidebar.error(f"로컬 파일 로드 실패: {e}")

    if uploaded is not None:
        # Avoid reading / hashing 100MB+ bytes on every rerun. If the same upload is still present,
        # re-use the already-parsed DataFrame from session_state.
        prev_name = st.session_state.get("upload_name")
        prev_size = st.session_state.get("upload_size")
        prev_df = st.session_state.get("raw_df")
        if prev_name == uploaded.name and prev_size == getattr(uploaded, "size", None) and isinstance(prev_df, pd.DataFrame):
            st.sidebar.success(f"업로드 로드(캐시): {uploaded.name}")
            return prev_df

        with st.sidebar.status("업로드 파일 읽는 중...", expanded=False):
            b = uploaded.getvalue()
            # MD5는 100MB+ 파일에서 체감으로 느려질 수 있어 생략합니다.
            key = f"upload:{uploaded.name}:{len(b)}"
            df = load_data_from_upload(uploaded.name, b)
        st.session_state["data_key"] = key
        st.session_state["raw_df"] = df
        st.session_state["upload_name"] = uploaded.name
        st.session_state["upload_size"] = getattr(uploaded, "size", None)
        st.sidebar.success(f"업로드 로드: {uploaded.name}")
        return df

    return None


def _render_header(repetitive_count: int) -> None:
    st.markdown(
        """
<style>
  .mds-title { display:flex; align-items:baseline; gap:12px; margin-top:6px; }
  .mds-brand { font-size:44px; font-weight:900; letter-spacing:-0.5px; color:#e94560; }
  .mds-name  { font-size:44px; font-weight:900; letter-spacing:-0.5px; color:#ffffff; }
  .mds-sub   { margin-top:-10px; color:#a0a0a0; }
  .mds-card  { background:#1e2a3a; border:1px solid #2d3748; border-radius:14px; padding:14px 16px; }
  .mds-muted { color:#a0a0a0; }
  .mds-badge { display:inline-block; padding:2px 10px; border-radius:999px; font-weight:700; font-size:12px; }
  .mds-badge-ok { background:#103a24; color:#9ff5c6; border:1px solid #1c6b41; }
  .mds-badge-bad { background:#3a1018; color:#ffd0d7; border:1px solid #7b2a37; }
  .mds-pill { display:inline-block; padding:4px 10px; border-radius:999px; font-weight:800; font-size:12px;
              background:#16213e; color:#ffffff; border:1px solid #2d3748; }
  .mds-kv { display:grid; grid-template-columns: 140px 1fr; gap:10px 14px; margin-top:12px; }
  .mds-k { color:#a0a0a0; font-size:12px; text-transform:none; letter-spacing:0.2px; }
  .mds-v { color:#ffffff; font-weight:650; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
</style>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        f"""
<div class="mds-title">
  <div class="mds-brand">T'way</div>
  <div class="mds-name">MDS</div>
</div>
<div class="mds-sub">Maintenance Document System (정비규정 4.3.3.10)</div>
        """,
        unsafe_allow_html=True,
    )

    if repetitive_count <= 0:
        st.markdown('<div class="mds-card"><span class="mds-badge mds-badge-ok">✅ 반복결함 해당사항 없음</span></div>', unsafe_allow_html=True)
    else:
        st.markdown(
            f'<div class="mds-card"><span class="mds-badge mds-badge-bad">🔴 반복결함 {repetitive_count}건 검출</span>'
            '<div class="mds-muted" style="margin-top:6px;">고장탐구 Manual에 따라 선제 해결 필요 (상단 탭의 “반복결함”에서 확인)</div></div>',
            unsafe_allow_html=True,
        )


def _render_search_tab(df: pd.DataFrame, mds: pd.DataFrame) -> None:
    f: SearchFilters = st.session_state["filters"]

    st.subheader("검색")
    st.caption("AC Type / Reg / W/O / Description / Corrective Action / Column N / ATA / 기간으로 검색합니다.")

    # Filtering over 200k+ rows is expensive; do it only when (data, filters) change.
    data_key = st.session_state.get("data_key")
    filter_key = (data_key, f)
    cached_for = st.session_state.get("filtered_for")
    cached_df = st.session_state.get("filtered_df")

    if cached_df is None or cached_for != filter_key:
        with st.spinner("검색 결과 계산 중..."):
            filtered = apply_filters(mds, f)
        st.session_state["filtered_df"] = filtered
        st.session_state["filtered_for"] = filter_key
        st.session_state["current_result"] = filtered
    else:
        filtered = cached_df
        st.session_state["current_result"] = filtered

    st.metric("조회 결과(건)", f"{len(filtered):,}")

    export_cols = ["My_AC_Str", "My_Reg", "My_WO", "My_Order_No", "My_Desc", "My_Action", "My_Col_N", "My_ATA", "My_Ref_Doc", "My_Date"]
    for col in export_cols:
        if col not in filtered.columns:
            filtered[col] = ""

    show_max = st.slider("표시 최대 행 수", min_value=100, max_value=50000, value=5000, step=100)
    # User-facing table: only 핵심 4개 컬럼(날짜/기번/결함/조치)
    table_cols = ["My_Date", "My_Reg", "My_Desc", "My_Action"]
    for col in table_cols:
        if col not in filtered.columns:
            filtered[col] = ""

    # IMPORTANT: build the table view only for the displayed slice (avoid work on full filtered DF).
    slice_df = filtered.head(show_max)
    view = slice_df[table_cols].copy()
    if "Display_Date" in slice_df.columns:
        view["My_Date"] = slice_df["Display_Date"].astype(str)
    else:
        view["My_Date"] = view["My_Date"].apply(format_date)
    view = view.rename(
        columns={
            "My_Date": "날짜",
            "My_Reg": "기번",
            "My_Desc": "결함 내용",
            "My_Action": "조치내용",
        }
    )

    def _shorten(s: Any, n: int = 160) -> str:
        t = "" if s is None else str(s).replace("\n", " ").strip()
        if len(t) <= n:
            return t
        return t[: n - 1] + "…"

    # Selection: click a row to select, then open detail via a single button.
    # Streamlit tables don't support per-row buttons reliably, and LinkColumn opens a new tab,
    # which breaks upload/session state. This approach stays in the same session.
    slice_with_idx = slice_df.reset_index(drop=False).rename(columns={"index": "_row_index"})
    view["결함 내용"] = view["결함 내용"].map(_shorten)
    view["조치내용"] = view["조치내용"].map(_shorten)

    st.caption("행을 클릭해 선택한 다음, 아래 **상세보기** 버튼을 누르면 새창(팝업)으로 상세가 열립니다.")
    event = st.dataframe(
        view,
        width="stretch",
        hide_index=True,
        on_select="rerun",
        selection_mode="single-row",
        key="results_table",
    )

    selected_row_index: int | None = None
    try:
        sel = getattr(event, "selection", None)
        if sel and sel.rows:
            pos = int(sel.rows[0])
            selected_row_index = int(slice_with_idx.iloc[pos]["_row_index"])
    except Exception:  # noqa: BLE001
        selected_row_index = None

    if selected_row_index is not None:
        try:
            st.session_state["selected_row"] = filtered.loc[selected_row_index].to_dict()
        except Exception:  # noqa: BLE001
            st.session_state["selected_row"] = None

    row = st.session_state.get("selected_row")
    if isinstance(row, dict) and row:
        summary = f"{format_date(row.get('My_Date'))} | {row.get('My_Reg','')} | {str(row.get('My_Desc',''))[:60]}"
        st.markdown(f"<div class='mds-muted'>선택됨: {html.escape(summary)}</div>", unsafe_allow_html=True)
    else:
        st.markdown("<div class='mds-muted'>선택된 행 없음</div>", unsafe_allow_html=True)

    if st.button("상세보기", type="primary", disabled=not (isinstance(row, dict) and row)):
        _show_detail_dialog(row)  # type: ignore[arg-type]

    c1, c2 = st.columns([1, 1])
    with c1:
        prep = st.button("다운로드 파일 준비", type="primary")
    with c2:
        st.caption("속도 개선: 버튼을 누를 때만 Excel/CSV를 생성합니다.")

    # Preparing large downloads on every rerun is expensive; do it on-demand.
    if prep:
        with st.spinner("다운로드 파일 생성 중..."):
            st.session_state["download_xlsx"] = _download_excel_bytes(filtered, f)
            st.session_state["download_csv"] = filtered.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
            st.session_state["download_ready_for"] = (st.session_state.get("data_key"), f)

    ready_for = st.session_state.get("download_ready_for")
    if ready_for == (st.session_state.get("data_key"), f) and st.session_state.get("download_xlsx") and st.session_state.get("download_csv"):
        d1, d2 = st.columns([1, 1])
        with d1:
            st.download_button(
                "💾 결과 저장 (Excel)",
                data=st.session_state["download_xlsx"],
                file_name=f"MDS_검색결과_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        with d2:
            st.download_button(
                "⬇️ 결과 CSV 다운로드",
                data=st.session_state["download_csv"],
                file_name="MDS_검색결과.csv",
                mime="text/csv",
            )
    else:
        st.info("다운로드가 필요하면 **다운로드 파일 준비**를 눌러주세요.")

    # Keep the main tab focused on search & table.


@st.dialog("결함 상세", width="large")
def _show_detail_dialog(row_dict: dict[str, Any]) -> None:
    _render_detail_tab(row_dict)

def _render_detail_tab(row_dict: dict[str, Any]) -> None:
    def _v(k: str) -> str:
        v = row_dict.get(k, "")
        if k == "My_Date":
            return format_date(v)
        return "" if v is None else str(v)

    reg = html.escape(_v("My_Reg"))
    act = html.escape(_v("My_AC_Str"))
    ata = html.escape(_v("My_ATA"))
    date = html.escape(_v("My_Date"))
    wo = html.escape(_v("My_WO"))
    order_no = html.escape(_v("My_Order_No"))
    col_n = html.escape(_v("My_Col_N"))
    ref_doc = html.escape(_v("My_Ref_Doc"))

    st.markdown(
        f"""
<div class="mds-card">
  <div style="display:flex; flex-wrap:wrap; gap:8px; align-items:center;">
    <span class="mds-pill">기번: {reg or '-'}</span>
    <span class="mds-pill">기종: {act or '-'}</span>
    <span class="mds-pill">ATA: {ata or '-'}</span>
    <span class="mds-pill">날짜: {date or '-'}</span>
  </div>
  <div class="mds-kv">
    <div class="mds-k">Tway W/O</div><div class="mds-v" title="{wo}">{wo or '-'}</div>
    <div class="mds-k">Order No.</div><div class="mds-v" title="{order_no}">{order_no or '-'}</div>
    <div class="mds-k">Col N</div><div class="mds-v" title="{col_n}">{col_n or '-'}</div>
    <div class="mds-k">Ref Doc</div><div class="mds-v" title="{ref_doc}">{ref_doc or '-'}</div>
  </div>
</div>
        """,
        unsafe_allow_html=True,
    )

    left, right = st.columns([1, 1])
    with left:
        st.markdown("**결함 내용(Description)**")
        st.text_area(
            label="결함 내용",
            value=_v("My_Desc") or "-",
            height=240,
            label_visibility="collapsed",
            disabled=True,
        )
    with right:
        st.markdown("**조치내용(Corrective Action)**")
        st.text_area(
            label="조치내용",
            value=_v("My_Action") or "-",
            height=240,
            label_visibility="collapsed",
            disabled=True,
        )

    tab1, tab2 = st.tabs(["추가 정보", "원본(전체 컬럼)"])
    with tab1:
        candidates = [
            ("Order Status", "Order Status"),
            ("Order Type", "Order Type"),
            ("Work Center", "Work Center"),
            ("Reported Name", "Reported Name"),
            ("Reported by", "Reported by"),
            ("Noti. No.", "Noti. No."),
            ("Noti. Type", "Noti. Type"),
            ("Noti. Status", "Noti. Status"),
            ("Func. Loc.", "Func. Loc."),
            ("Equip.", "Equip."),
            ("MFG P/N", "MFG P/N"),
            ("MFG S/N", "MFG S/N"),
            ("Flight No.", "Flight No."),
            ("Ref.DOC", "Ref.DOC"),
            ("REF DOC", "REF DOC"),
            ("REF DOC No.", "REF DOC No."),
            ("REF Task Card No.", "REF Task Card No."),
        ]
        rows: list[dict[str, str]] = []
        for label, key in candidates:
            val = row_dict.get(key, "")
            if val is None:
                continue
            sval = str(val).strip()
            if not sval:
                continue
            rows.append({"항목": label, "값": sval})

        if rows:
            st.dataframe(pd.DataFrame(rows), width="stretch", hide_index=True)
        else:
            st.info("추가로 표시할 값이 없습니다.")

    with tab2:
        st.json(row_dict)


def _render_repetitive_tab(repetitive: list[dict[str, Any]]) -> None:
    st.subheader("🔴 반복결함 검토 (정비규정 4.3.3.10)")
    st.markdown(
        """
- 정의: 동일항공기의 특정 계통 내에서 발생한 동일 성질의 결함이 모기지 경유 **1주 3회 / 2주 5회 / 4주 8회** 이상 발생
- 제외: ATA 32/33/34, 좌석위치별결함(ATA25), MEL/CDL REPETITIVE 점검, 외관 결함
        """
    )

    if not repetitive:
        st.success("반복결함 해당사항 없음")
        return

    rep_df = pd.DataFrame(repetitive).drop(columns=["dates"], errors="ignore")
    st.metric("검출 건수", f"{len(rep_df):,}")
    st.dataframe(rep_df, width="stretch")

    st.download_button(
        "📊 반복결함 Excel 저장",
        data=df_to_xlsx_bytes(rep_df, sheet_name="반복결함"),
        file_name=f"반복결함_정비규정4.3.3.10_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.divider()
    st.subheader("빠른 검색")
    st.caption("선택한 반복결함(Reg/ATA)로 검색 조건을 자동 세팅합니다. (기간은 4주로 설정)")

    options = [f"{i:03d} | {d['reg']} | ATA {d['ata']} | {d['criteria']} | {d['symptom']}" for i, d in enumerate(repetitive)]
    chosen = st.selectbox("반복결함 선택", options=options, index=0)
    chosen_idx = int(chosen.split("|", 1)[0].strip())
    picked = repetitive[chosen_idx]

    if st.button("이 조건으로 검색 세팅"):
        today = datetime.now().date()
        st.session_state["filters"] = SearchFilters(
            **{
                **st.session_state["filters"].__dict__,
                "reg": str(picked["reg"]),
                "ata": str(picked["ata"]),
                "date_from": (today - timedelta(days=28)).strftime("%Y-%m-%d"),
                "date_to": today.strftime("%Y-%m-%d"),
            }
        )
        st.success("검색 조건을 업데이트했습니다. 상단의 '검색' 탭에서 확인하세요.")


def _render_statistics_tab(mds: pd.DataFrame) -> None:
    st.subheader("📊 통계 분석 (ATA 챕터별)")
    if not MATPLOTLIB_AVAILABLE:
        st.warning("matplotlib이 설치되어 있지 않아 통계를 표시할 수 없습니다.")
        return

    # IMPORTANT: Tabs re-run the whole script; compute only on demand and cache by key.
    current = st.session_state.get("current_result")
    data_source = st.radio("데이터 범위", options=["현재 검색결과", "전체 데이터"], horizontal=True)
    period = st.radio("기간", options=["전체", "1주", "2주", "4주", "12주"], horizontal=True)

    # Use a compute button to avoid heavy work on every rerun.
    c1, c2 = st.columns([1, 2])
    with c1:
        run = st.button("통계 그리기", type="primary")
    with c2:
        st.caption("속도 개선을 위해 버튼을 눌렀을 때만 통계를 계산합니다.")

    stats_key = (st.session_state.get("data_key"), data_source, period)
    if run:
        st.session_state["stats_ready"] = True
        st.session_state["stats_key"] = stats_key

    if not st.session_state.get("stats_ready") or st.session_state.get("stats_key") != stats_key:
        st.info("위에서 조건을 정한 뒤 **통계 그리기**를 눌러주세요.")
        return

    # Base dataset for stats
    if data_source == "현재 검색결과" and isinstance(current, pd.DataFrame) and not current.empty:
        df = current
    else:
        df = mds

    if period != "전체":
        days = {"1주": 7, "2주": 14, "4주": 28, "12주": 84}[period]
        cutoff = datetime.now() - timedelta(days=days)
        df = df[(df["Parsed_Date"].notna()) & (df["Parsed_Date"] >= cutoff)]

    if df.empty:
        st.info("통계 대상 데이터가 없습니다.")
        return

    # Cache base stats
    stats_cache = st.session_state.setdefault("stats_cache", {})
    if stats_key not in stats_cache:
        with st.spinner("통계 계산 중..."):
            ata_counts = df["My_ATA"].astype(str).value_counts()
            ac_counts = df["My_AC_Str"].astype(str).value_counts()
            reg_counts = df["My_Reg"].astype(str).value_counts()
            stats_cache[stats_key] = {
                "ata_counts": ata_counts,
                "ac_counts": ac_counts,
                "reg_counts": reg_counts,
                "total": int(len(df)),
                "unique_ac": int(df["My_AC_Str"].nunique()),
            }

    base = stats_cache[stats_key]
    ata_counts: pd.Series = base["ata_counts"]
    ac_counts: pd.Series = base["ac_counts"]
    reg_counts: pd.Series = base["reg_counts"]

    # KPI cards
    st.subheader("요약")
    top_ata = ata_counts.index[0] if not ata_counts.empty else "-"
    top_ac = ac_counts.index[0] if not ac_counts.empty else "-"
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("총 건수", f"{base['total']:,}")
    k2.metric("고유 기종 수", f"{base['unique_ac']:,}")
    k3.metric("Top ATA", f"{top_ata}")
    k4.metric("Top 기종", f"{top_ac}")

    st.divider()
    st.subheader("기종×ATA Top10 + Pareto")
    ac_options = ["전체"] + ac_counts.index.tolist()
    sel_ac = st.selectbox("기종 선택", options=ac_options, index=0)
    if sel_ac == "전체":
        df_ac = df
    else:
        df_ac = df[df["My_AC_Str"].astype(str) == sel_ac]

    if df_ac.empty:
        st.info("선택한 기종에 대한 데이터가 없습니다.")
    else:
        ata_top = df_ac["My_ATA"].astype(str).value_counts().head(10)
        fig = _plot_pareto(ata_top, title=f"기종 {sel_ac} - ATA Top10", xlabel="ATA")
        st.pyplot(fig, width="stretch")

    st.divider()
    st.subheader("ATA별 결함 키워드 Top10 + Pareto")
    ata_options = ata_counts.index.tolist()
    if ata_options:
        sel_ata = st.selectbox("ATA 선택", options=ata_options, index=0)
        keyword_cache = st.session_state.setdefault("keyword_cache", {})
        kw_key = ("ata", stats_key, sel_ata)
        if kw_key not in keyword_cache:
            with st.spinner("키워드 추출 중..."):
                df_ata = df[df["My_ATA"].astype(str) == sel_ata]
                tokens = extract_keywords(df_ata["My_Desc"])
                keyword_cache[kw_key] = tokens.value_counts().head(10)
        kw_counts: pd.Series = keyword_cache.get(kw_key, pd.Series(dtype=int))
        if kw_counts.empty:
            st.info("해당 ATA에서 추출할 키워드가 없습니다.")
        else:
            fig = _plot_pareto(kw_counts, title=f"ATA {sel_ata} - 결함 키워드 Top10", xlabel="키워드")
            st.pyplot(fig, width="stretch")
    else:
        st.info("ATA 데이터가 없습니다.")

    st.divider()
    st.subheader("기번(선택형) 심화")
    reg_options = reg_counts.index[:200].tolist()
    selected_regs = st.multiselect("기번 선택 (최대 3개 권장)", options=reg_options, default=[])
    if selected_regs:
        selected_regs = selected_regs[:3]
        for reg in selected_regs:
            st.markdown(f"**기번: {reg}**")
            df_reg = df[df["My_Reg"].astype(str) == reg]
            if df_reg.empty:
                st.info("해당 기번 데이터가 없습니다.")
                continue

            # ATA 분포 (막대)
            ata_reg = df_reg["My_ATA"].astype(str).value_counts().head(10)
            fig = plt.figure(figsize=(10, 4))
            ax = fig.add_subplot(1, 1, 1)
            ata_reg.plot(kind="bar", ax=ax, color="#6ab0ff")
            ax.set_title(f"{reg} - ATA 분포(Top10)")
            ax.set_xlabel("ATA")
            ax.set_ylabel("건수")
            fig.tight_layout()
            st.pyplot(fig, width="stretch")

            # 결함 키워드 Top10
            kw_key = ("reg", stats_key, reg)
            keyword_cache = st.session_state.setdefault("keyword_cache", {})
            if kw_key not in keyword_cache:
                with st.spinner("키워드 추출 중..."):
                    tokens = extract_keywords(df_reg["My_Desc"])
                    keyword_cache[kw_key] = tokens.value_counts().head(10)
            kw_counts = keyword_cache.get(kw_key, pd.Series(dtype=int))
            if kw_counts.empty:
                st.info("키워드가 없습니다.")
            else:
                fig = _plot_pareto(kw_counts, title=f"{reg} - 결함 키워드 Top10", xlabel="키워드")
                st.pyplot(fig, width="stretch")
    else:
        st.caption("기번을 선택하면 상세 통계가 표시됩니다.")


def _render_gemini_tab() -> None:
    st.subheader("✨ Gemini 2.0")

    if not GEMINI_AVAILABLE:
        st.warning("`google-generativeai`가 설치되어 있지 않습니다. requirements.txt에 추가되어 있으니 배포 환경에서는 설치됩니다.")
        return

    st.caption("이 탭은 ‘T’way MDS 웹’ 안에서만 답변하도록(범위 제한) 설계되어 있습니다.")

    # Prefer Streamlit secrets. Allow fallback input for local-only runs.
    api_key = ""
    try:
        api_key = st.secrets.get("GEMINI_API_KEY", "")
    except Exception:  # noqa: BLE001
        api_key = ""
    if not api_key:
        api_key = st.sidebar.text_input("Gemini API Key (로컬 테스트용)", type="password", value="")

    model_name = st.selectbox("모델", options=["gemini-2.0-flash", "gemini-1.5-pro", "gemini-1.5-flash"], index=0)
    st.session_state["gemini_model"] = model_name

    if not api_key:
        st.info("Gemini API Key가 없어서 채팅을 실행할 수 없습니다. (Streamlit Cloud에서는 Secrets에 `GEMINI_API_KEY` 설정)")
        return

    genai.configure(api_key=api_key)

    if not st.session_state["gemini_messages"]:
        st.session_state["gemini_messages"] = [
            {"role": "assistant", "content": "안녕하세요! T'way MDS AI입니다.\n정비규정 4.3.3.10 반복결함 관련 질문이나 정비 관련 문의사항을 말씀해주세요."}
        ]

    for m in st.session_state["gemini_messages"]:
        with st.chat_message(m["role"]):
            st.write(m["content"])

    use_selected = st.toggle("선택한 결함(상세 보기)을 컨텍스트로 사용", value=True)
    use_filters = st.toggle("현재 검색조건/통계 요약을 컨텍스트로 사용", value=True)

    user_msg = st.chat_input("메시지를 입력하세요")
    if not user_msg:
        return

    st.session_state["gemini_messages"].append({"role": "user", "content": user_msg})
    with st.chat_message("user"):
        st.write(user_msg)

    with st.chat_message("assistant"):
        with st.spinner("Gemini 응답 생성 중..."):
            try:
                # Build a strict system instruction so the bot stays inside the MDS app.
                filters: SearchFilters = st.session_state.get("filters")
                current: pd.DataFrame | None = st.session_state.get("current_result")
                selected_row = st.session_state.get("selected_row") if use_selected else None

                ctx_lines: list[str] = []
                if use_filters and isinstance(filters, SearchFilters):
                    ctx_lines.append("## 현재 검색조건")
                    ctx_lines.append(
                        f"- AC Type: {filters.actype or '전체'}\n"
                        f"- Reg No: {filters.reg or '전체'}\n"
                        f"- Work Order: {filters.wo or '전체'}\n"
                        f"- ATA: {filters.ata or '전체'}\n"
                        f"- 기간: {filters.date_from or '-'} ~ {filters.date_to or '-'}"
                    )
                if use_filters and isinstance(current, pd.DataFrame):
                    ctx_lines.append(f"- 현재 검색결과 건수: {len(current):,}")

                if selected_row and isinstance(selected_row, dict):
                    ctx_lines.append("## 선택한 결함(행) 요약")
                    # Keep it compact: key fields only.
                    def _sv(k: str) -> str:
                        v = selected_row.get(k, "")
                        if k == "My_Date":
                            return format_date(v)
                        return "" if v is None else str(v)

                    ctx_lines.append(
                        "\n".join(
                            [
                                f"- Reg: {_sv('My_Reg')}",
                                f"- AC Type: {_sv('My_AC_Str')}",
                                f"- ATA: {_sv('My_ATA')}",
                                f"- Date: {_sv('My_Date')}",
                                f"- Description: {_sv('My_Desc')[:400]}",
                                f"- Corrective Action: {_sv('My_Action')[:400]}",
                            ]
                        )
                    )

                # Lightweight grounding: pull up to 8 similar rows from current_result by keyword overlap.
                if isinstance(current, pd.DataFrame) and not current.empty:
                    q = user_msg.strip()
                    tokens = [t for t in re.split(r"[^0-9A-Za-z가-힣]+", q) if len(t) >= 3][:12]
                    if tokens:
                        score = pd.Series([0] * len(current), index=current.index)
                        for t in tokens:
                            if "My_Desc" in current.columns:
                                score += current["My_Desc"].astype(str).str.contains(re.escape(t), case=False, na=False).astype(int)
                            if "My_Action" in current.columns:
                                score += current["My_Action"].astype(str).str.contains(re.escape(t), case=False, na=False).astype(int)
                        top = current.loc[score.sort_values(ascending=False).head(8).index]
                        top = top[["My_Reg", "My_AC_Str", "My_ATA", "My_Date", "My_Desc", "My_Action"]].copy()
                        top["My_Date"] = top["My_Date"].apply(format_date)
                        ctx_lines.append("## 참고(유사 이력 샘플, 최대 8건)")
                        ctx_lines.append(top.to_markdown(index=False))

                system_instruction = (
                    "당신은 ‘T’way MDS 웹’ 전용 AI 어시스턴트입니다.\n"
                    "반드시 아래 범위에서만 답하세요:\n"
                    "- MDS 이력 조회 화면에서 사용자가 보고 있는 데이터/필터/선택행/반복결함 규정(4.3.3.10) 관련 질문\n"
                    "- 검색조건 설계/해석, 데이터 컬럼 의미 추정(근거가 있으면), 반복결함 기준 설명, 보고서 요약\n\n"
                    "금지:\n"
                    "- 일반 잡담/광범위한 지식 설명(항공/정비 일반론 포함)\n"
                    "- 법/의학/재무 등 MDS와 무관한 조언\n\n"
                    "규칙:\n"
                    "- 답변은 한국어로, 핵심만 5~10줄로 요약\n"
                    "- 필요한 경우에만 짧게 근거(어떤 컬럼/어떤 값 기반인지) 표시\n"
                    "- 주어진 컨텍스트(필터/선택행/유사샘플) 밖의 사실은 단정하지 말 것\n"
                    "- 질문이 범위를 벗어나면: '이 MDS 웹에서는 해당 범위를 다루지 않습니다'라고 답하고, MDS 관련으로 다시 질문 유도\n"
                )

                prompt = "\n\n".join([system_instruction, *ctx_lines, "## 사용자 질문", user_msg])
                model = genai.GenerativeModel(model_name, system_instruction=system_instruction)
                resp = model.generate_content(prompt)
                text = getattr(resp, "text", None) or str(resp)
            except Exception as e:  # noqa: BLE001
                text = f"⚠️ 오류: {e}"
        st.write(text)

    st.session_state["gemini_messages"].append({"role": "assistant", "content": text})


def main() -> None:
    st.set_page_config(page_title=APP_TITLE, layout="wide")
    _init_state()
    require_login()

    _render_sidebar_filters()
    df = _load_ui()

    if df is None:
        st.title(APP_TITLE)
        st.info("사이드바에서 파일을 업로드하거나(권장), 로컬 파일 경로를 입력해 주세요.")
        return

    # Compute once per loaded file. Streamlit reruns on almost any interaction,
    # so doing heavy work unconditionally makes the app *feel* unresponsive.
    key = st.session_state.get("data_key")
    if key is None:
        key = f"fallback:{len(df)}:{len(df.columns)}"
        st.session_state["data_key"] = key

    if st.session_state.get("mds_df") is None or st.session_state.get("repetitive") is None or st.session_state.get("_computed_for") != key:
        with st.status("데이터 준비 중... (처음 1회만 시간이 걸릴 수 있어요)", expanded=False) as status:
            status.update(label="1/3 컬럼 매핑 생성 중...")
            mds = build_mds_columns(df)
            status.update(label="2/3 반복결함 분석 중...")
            repetitive = analyze_repetitive_defects(mds)
            status.update(label="3/3 화면 준비 중...")
            st.session_state["mds_df"] = mds
            st.session_state["repetitive"] = repetitive
            st.session_state["_computed_for"] = key
            # Reset current_result when the underlying data changes.
            st.session_state["current_result"] = None
            st.session_state["filtered_df"] = None
            st.session_state["filtered_for"] = None
            st.session_state["download_ready_for"] = None
            st.session_state["download_xlsx"] = None
            st.session_state["download_csv"] = None
            status.update(label="완료", state="complete")
    else:
        mds = st.session_state["mds_df"]
        repetitive = st.session_state["repetitive"]

    setup_korean_font()

    _render_header(len(repetitive))

    tab_search, tab_rep, tab_stats, tab_gemini = st.tabs(["검색", "반복결함", "통계 분석", "Gemini 2.0"])

    with tab_search:
        _render_search_tab(df, mds)
    with tab_rep:
        _render_repetitive_tab(repetitive)
    with tab_stats:
        _render_statistics_tab(mds)
    with tab_gemini:
        _render_gemini_tab()


if __name__ == "__main__":
    main()
